import openpyxl
from datetime import datetime
import os

# Função para criar um novo arquivo de finanças
def create_finance_file(sheet="financas.xlsx"):
    if os.path.exists(sheet):
        os.remove(sheet)  # Excluir o arquivo existente
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Despesas"
    ws.append(["Data", "Valor", "Categoria", "Saldo"])  # Adicionar cabeçalho para o saldo
    wb.save(sheet)
    print(f"Arquivo '{sheet}' criado com sucesso.")

# Função para registrar uma nova despesa na planilha e atualizar o saldo
def record_expense(value, category, sheet="financas.xlsx"):
    wb = openpyxl.load_workbook(sheet)
    ws = wb.active
    current_balance = ws.cell(row=ws.max_row, column=4).value  # Obter o saldo atual da última linha da coluna de saldo
    
    # Verificar se o saldo_atual é uma string (pode ser o título da coluna)
    if isinstance(current_balance, str):
        print("Sem saldo definido. Impossível registrar a despesa.")
        return
    
    # Verificar se há saldo suficiente para a despesa
    if current_balance is None or current_balance < value:
        print("Saldo insuficiente. Impossível registrar a despesa.")
        return
    
    # Registrar a despesa e atualizar o saldo
    ws.append([datetime.now().strftime("%Y-%m-%d"), f"-{value}", category, current_balance - value])  # Atualizar o saldo após despesa
    wb.save(sheet)
    print(f"Despesa registrada com sucesso. - R$ {value}")


# Função para verificar o saldo total na planilha
def check_balance(sheet="financas.xlsx"):
    wb = openpyxl.load_workbook(sheet)
    ws = wb.active
    current_balance = ws.cell(row=ws.max_row, column=4).value  # Obter o saldo da última linha da coluna de saldo
    
    # Verificar se o saldo_atual é uma string (pode ser o título da coluna)
    if isinstance(current_balance, str):
        print("Sem saldo presente definido.")
        return
    
    print(f"Seu saldo atual é de R${current_balance:.2f}.")

# Função para definir o orçamento mensal na planilha
def set_budget(value, sheet="financas.xlsx"):
    wb = openpyxl.load_workbook(sheet)
    ws = wb.active
    current_balance = None
    for row in ws.iter_rows(min_row=2, max_col=4, max_row=ws.max_row, values_only=True):
        if row[3] is not None:
            current_balance = row[3]
            break
    new_balance = float(current_balance) + value if current_balance is not None and float(current_balance) > 0 else value
    # Adicionar nova linha com o orçamento definido
    ws.append([datetime.now().strftime("%Y-%m-%d"), f"+{value}", "orçamento", new_balance])
    wb.save(sheet)
    print(f"Orçamento mensal definido com sucesso. + R$ {value}")

# Função para visualizar o relatório de gastos do mês atual na planilha
def view_report(sheet="financas.xlsx"):
    # verificar se o arquivo existe
    if not os.path.exists(sheet):
        print("O arquivo de finanças não foi encontrado.")
        return
    
    try:
        os.startfile(sheet)
        print("Relatório de gastos aberto com sucesso.")
    except Exception as e:
        print(f"Erro ao abrir o arquivo: {e}")

# Função para atuar sobre as finanças com base na ação solicitada
def act_on_finance(action, params, _):
    if action == "registrar" and 'despesa' in params: 
        value = float(params[1])  # Converter para float
        record_expense(value, params[2])
    elif action == "verificar" and 'saldo' in params: 
        check_balance()
    elif action == "definir" and 'orçamento' in params: 
        value = float(params[1])
        set_budget(value)
    elif action == "visualizar" and 'relatório' in params: 
        view_report()
    else:
        print("Ação financeira não reconhecida.")